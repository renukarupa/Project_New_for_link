import pytest
import openpyxl
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from datetime import datetime
from login_page import LoginPage
from pim_page import PimPage

EXCEL_PATH = 'test_data.xlsx'

def get_test_data():
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    test_data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        test_data.append({
            "test_id": row[0],
            "username": row[1],
            "password": row[2],
            "tester_name": row[3]
        })
    workbook.close()
    return test_data

def write_test_result(test_id, tester_name, result):
    workbook = openpyxl.load_workbook(EXCEL_PATH)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == test_id:
            row[4].value = result
            row[5].value = datetime.now().date()
            row[6].value = datetime.now().time()
            break
    workbook.save(EXCEL_PATH)
    workbook.close()

@pytest.mark.parametrize("data", get_test_data())
def test_login(data):
    driver = webdriver.Chrome()
    driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
    login_page = LoginPage(driver)

    login_page.enter_username(data["username"])
    login_page.enter_password(data["password"])
    login_page.click_login()

    try:
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.ID, "menu_dashboard_index"))
        )
        pim_page = PimPage(driver)
        pim_page.navigate_to_pim()
        pim_page.click_employee_list()
        pim_page.delete_first_employee()

        result = "Passed"
    except Exception as e:
        result = f"Failed: {e}"
    finally:
        write_test_result(data["test_id"], data["tester_name"], result)
        driver.quit()
